import time
import threading
import pandas as pd
import numpy as np
from scipy.signal import find_peaks
import keyboard  # Klavye girişlerini izlemek için
from openpyxl import Workbook

# Seri port haberleşmesi için alternatif
try:
    import serial
    Serial = serial.Serial
except ImportError:
    print("PySerial kütüphanesi bulunamadı. Alternatif yöntemler kullanılacak.")
    Serial = None

def initialize_serial(port, baudrate):
    if Serial is not None:
        try:
            ser = Serial(port, baudrate)
            time.sleep(2)  # Bağlantının kurulması için kısa bir gecikme
            return ser
        except Exception as e:
            print(f"Seri bağlantı başlatılamadı: {e}")
            return None
    else:
        print("Seri haberleşme desteklenmiyor.")
        return None

# Seri port ayarları
ser = initialize_serial('COM7', 115200)
if ser is None:
    print("Seri bağlantı başlatılamadı. Çıkılıyor...")
    exit()

# Excel dosyası oluşturma
wb = Workbook()
ws = wb.active
ws.title = "EKG Data"

# Geçici tampon (buffer) tanımlama
buffer = []
batch_size_normal = 100  # Normal modda her satır için 100 veri
batch_size_tachycardia = 250  # Taşikardi modunda tampon boyutu
batch_size = batch_size_normal  # Varsayılan tampon boyutu
current_row = 1  # İlk yazılacak satır (başlangıç)
skip_count = 3  # Atlanacak veri grubu sayısı

# Taşikardi modu
tachycardia_mode = False

# Excel kaydetme işlemi için bir iş parçacığı tanımlama
save_lock = threading.Lock()

def save_to_excel_periodically():
    while True:
        time.sleep(5)  # 5 saniyede bir kaydet
        with save_lock:
            try:
                wb.save("megadeneme.xlsx")
                print("Dosya başarıyla kaydedildi.")
            except PermissionError:
                print("Dosya açık olduğu için kaydedilemedi.")

# Taşikardi sinyali oluşturma
def generate_tachycardia_signal(signal, speed_factor=2.5, amplitude_factor=1.2, peak_amplification=1.5):
    original_indices = np.linspace(0, len(signal) - 1, len(signal))
    new_length = int(len(signal) / speed_factor)
    new_indices = np.linspace(0, len(signal) - 1, new_length)
    accelerated_signal = np.interp(new_indices, original_indices, signal)
    accelerated_signal *= amplitude_factor

    peaks, _ = find_peaks(accelerated_signal, height=np.mean(accelerated_signal))
    enhanced_signal = accelerated_signal.copy()
    for peak in peaks:
        enhanced_signal[peak] *= peak_amplification

    return enhanced_signal

# Kaydetme işlemini başlatan iş parçacığı
save_thread = threading.Thread(target=save_to_excel_periodically, daemon=True)
save_thread.start()

try:
    while True:
        if keyboard.is_pressed('s'):
            tachycardia_mode = not tachycardia_mode
            batch_size = batch_size_tachycardia if tachycardia_mode else batch_size_normal
            mode = "Taşikardi" if tachycardia_mode else "Normal"
            print(f"Mod değiştirildi: {mode}")
            time.sleep(1)  # Anahtar tekrar basılmasını önlemek için bekle

        if ser and ser.in_waiting > 0:
            raw_data = ser.readline()
            try:
                line = raw_data.decode('utf-8').strip()
                signal_value = float(line)
            except (UnicodeDecodeError, ValueError):
                print(f"Decoding error or invalid value: {raw_data}")
                continue

            print(f"{time.strftime('%H:%M:%S')} - {signal_value}")
            buffer.append(signal_value)

            if len(buffer) >= batch_size:
                if tachycardia_mode:
                    processed_signal = generate_tachycardia_signal(buffer[:batch_size])
                else:
                    processed_signal = buffer[:batch_size]

                if len(processed_signal) < 100:
                    print("Taşikardi modunda elde edilen veri 100'den az, işlem yapılamadı.")
                    buffer.clear()
                    continue

                if skip_count == 0:
                    with save_lock:
                        for i, value in enumerate(processed_signal, start=1):
                            ws.cell(row=current_row, column=i, value=value)
                        current_row += 1
                        skip_count = 3
                else:
                    skip_count -= 1

                del buffer[:batch_size]
except KeyboardInterrupt:
    print("Veri kaydı durduruldu.")
finally:
    if len(buffer) > 0:
        with save_lock:
            for i, value in enumerate(buffer, start=1):
                ws.cell(row=current_row, column=i, value=value)
    if ser:
        ser.close()
    wb.save("megadeneme.xlsx")
